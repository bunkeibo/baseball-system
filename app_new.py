import streamlit as st
import pandas as pd
import openpyxl
import plotly.express as px
import datetime

st.set_page_config(page_title="チーム野球成績管理システム", layout="wide")

st.title("⚾ チーム・個人成績チェックシステム")
st.write("Excelデータから自動連携された最新の成績を表示しています。")

# --- 1. データの読み込み関数 ---
@st.cache_data(ttl=5)
def load_excel_data():
    file = "baseball.xlsx"
    sheet_out = "統計"
    
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb[sheet_out]
    
    yashu_blocks = {}
    toushu_blocks = {}
    
    curr_month = None
    curr_start = None
    
    # Excelの全行を走査してブロックごとにデータを切り出す
    for r_idx in range(1, ws.max_row + 1):
        val = ws.cell(r_idx, 1).value
        
        # タイトル行（■ 通算、■ 2026-04 成績 など）を検知
        if isinstance(val, str) and val.startswith("■"):
            curr_month = val.replace("■ ", "").replace(" 成績", "").strip()
            continue
            
        # ヘッダー行「番号」を検知
        if val == "番号":
            curr_start = r_idx + 1
            continue
            
        # データの終わり（空行、またはシートの最下部）を検知した時の処理
        if curr_start and (val is None or r_idx == ws.max_row):
            data_end = r_idx if (r_idx == ws.max_row and val is not None) else r_idx - 1
            
            if curr_month and curr_start <= data_end:
                # --- 野手データの抽出 (A~O列: 1~15番目) ---
                y_data = []
                for r in range(curr_start, data_end + 1):
                    no_val = ws.cell(r, 1).value
                    if no_val is not None:
                        r_vals = [ws.cell(r, c).value for c in range(1, 16)]
                        y_data.append(r_vals)
                if y_data:
                    y_cols = ["番号","name","打数","安打","塁打数","四死球","三振","企図","成功","犠飛","打点","打率","長打率","出塁率","OPS"]
                    yashu_blocks[curr_month] = pd.DataFrame(y_data, columns=y_cols)
                
                # --- 投手データの抽出 (U~AF列: 21~32番目) ---
                p_data = []
                for r in range(curr_start, data_end + 1):
                    no_val_p = ws.cell(r, 21).value
                    if no_val_p is not None:
                        r_vals = [ws.cell(r, c).value for c in range(21, 33)]
                        p_data.append(r_vals)
                if p_data:
                    p_cols = ["番号","name","投球回","被安打","与四死球","奪三振","自責点","防御率","被打率","与球率","奪三振率","WHIP"]
                    toushu_blocks[curr_month] = pd.DataFrame(p_data, columns=p_cols)
                    
            curr_start = None
            
    return yashu_blocks, toushu_blocks

try:
    yashu_months, toushu_months = load_excel_data()
except Exception as e:
    st.error(f"Excelの読み込みに失敗しました。ファイルが閉じられているか確認してください。: {e}")
    st.stop()

# 利用可能な「月」のリストを作成（通算は除く）
months_only = sorted([k for k in yashu_months.keys() if "通算" not in k])

# --- 2. 画面UIの構築 (サイドバーで条件選択) ---
st.sidebar.header("🔍 表示条件設定")

mode = st.sidebar.radio(
    "表示モードを選択してください：",
    ["🏆 すべての通算成績", "🗓️ 今月の成績", "🌸 春季リーグ戦（4月・5月）"]
)

# モードに応じたデータの取得・統合処理
def get_filtered_data(mode):
    if mode == "🏆 すべての通算成績":
        # エクセルの「通算（全試合合計）」または「通算」という名前のブロックを取得
        k = "通算（全試合合計）" if "通算（全試合合計）" in yashu_months else "通算"
        return yashu_months.get(k, pd.DataFrame()), toushu_months.get(k, pd.DataFrame())
    
    elif mode == "🗓️ 今月の成績":
        # 今日の日付から月（2026-05など）を自動取得
        current_m_str = datetime.datetime.now().strftime("%Y-%m")
        if current_m_str not in yashu_months and months_only:
            current_m_str = months_only[-1] # データになければ直近の最新月
        return yashu_months.get(current_m_str, pd.DataFrame()), toushu_months.get(current_m_str, pd.DataFrame())
        
    elif mode == "🌸 春季リーグ戦（4月・5月）":
        spring_months = [m for m in months_only if m.endswith("-04") or m.endswith("-05")]
        
        # 野手4,5月合計
        y_list = [yashu_months[m] for m in spring_months if m in yashu_months]
        if y_list:
            df_s_y = pd.concat(y_list).groupby(["番号", "name"], as_index=False).sum()
            df_s_y["打率"] = (df_s_y["安打"] / df_s_y["打数"]).fillna(0)
            df_s_y["長打率"] = (df_s_y["塁打数"] / df_s_y["打数"]).fillna(0)
            df_s_y["出塁率"] = ((df_s_y["安打"] + df_s_y["四死球"]) / (df_s_y["打数"] + df_s_y["四死球"] + df_s_y["犠飛"])).fillna(0)
            df_s_y["OPS"] = df_s_y["出塁率"] + df_s_y["長打率"]
        else: df_s_y = pd.DataFrame()
            
        # 投手4,5月合計
        p_list = [toushu_months[m] for m in spring_months if m in toushu_months]
        if p_list:
            df_s_p = pd.concat(p_list).groupby(["番号", "name"], as_index=False).sum()
            df_s_p["防御率"] = ((df_s_p["自責点"] * 9) / df_s_p["投球回"]).fillna(0)
            df_s_p["WHIP"] = ((df_s_p["被安打"] + df_s_p["与四死球"]) / df_s_p["投球回"]).fillna(0)
            df_s_p["被打率"] = (df_s_p["被安打"] / df_s_p["投球回"]).fillna(0)
            df_s_p["与球率"] = ((df_s_p["与四死球"] * 9) / df_s_p["投球回"]).fillna(0)
            df_s_p["奪三振率"] = ((df_s_p["奪三振"] * 9) / df_s_p["投球回"]).fillna(0)
        else: df_s_p = pd.DataFrame()
            
        return df_s_y, df_s_p

df_active_y, df_active_p = get_filtered_data(mode)

# 選手・対象リストの作成（チーム全体を先頭にする）
all_players = set()
if not df_active_y.empty: all_players.update(df_active_y["name"].tolist())
if not df_active_p.empty: all_players.update(df_active_p["name"].tolist())
all_players = sorted(list(all_players))

selected_target = st.sidebar.selectbox(
    "👤 選手・対象を選択してください：",
    ["👥 チーム全体"] + all_players
)

# --- 3. メイン画面の描画 ---
st.header(f"{mode} ── 表示中: {selected_target}")

if selected_target == "👥 チーム全体":
    # === ① チーム全体の成績を一気に見る画面 ===
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("🔹 野手成績（ランキング順）")
        if not df_active_y.empty:
            df_show_y = df_active_y.sort_values(by="OPS", ascending=False)
            st.dataframe(
                df_show_y.style.background_gradient(cmap="YlGn", subset=["打率", "長打率", "出塁率", "OPS"])
                              .format({c: "{:.3f}" for c in ["打率", "長打率", "出塁率", "OPS"]}), 
                height=600, use_container_width=True
            )
        else:
            st.info("対象の野手データがありません。")
            
    with col2:
        st.subheader("🔸 投手成績（防御率順）")
        if not df_active_p.empty:
            df_show_p = df_active_p.sort_values(by="防御率", ascending=True)
            st.dataframe(
                df_show_p.style.background_gradient(cmap="OrRd_r", subset=["防御率", "WHIP"])
                              .format({c: "{:.2f}" for c in ["防御率", "WHIP", "与球率", "奪三振率"]})
                              .format({"被打率": "{:.3f}"}), 
                height=600, use_container_width=True
            )
        else:
            st.info("対象の投手データがありません。")

else:
    # === ② 個人成績 ＋ レーダーチャート画面 ===
    p_y = df_active_y[df_active_y["name"] == selected_target] if not df_active_y.empty else pd.DataFrame()
    p_p = df_active_p[df_active_p["name"] == selected_target] if not df_active_p.empty else pd.DataFrame()
    
    if p_y.empty and p_p.empty:
        st.warning("選択された期間にこの選手のデータはありません。")
    else:
        col_data, col_graph = st.columns([4, 3])
        
        with col_data:
            if not p_y.empty:
                st.subheader("⚾ 野手個人スタッツ")
                st.dataframe(p_y.style.format({c: "{:.3f}" for c in ["打率", "長打率", "出塁率", "OPS"]}), use_container_width=True)
                
            if not p_p.empty:
                st.subheader("🥎 投手個人スタッツ")
                st.dataframe(p_p.style.format({c: "{:.2f}" for c in ["防御率", "WHIP", "与球率", "奪三振率"]}).format({"被打率": "{:.3f}"}), use_container_width=True)
                
        with col_graph:
            st.subheader("📊 選手能力レーダーチャート")
            
            # --- 野手のレーダーチャート描画 ---
            if not p_y.empty:
                row = p_y.iloc[0]
                stats = {
                    "打率(ミート)": min(row["打率"] / 0.400, 1.0),
                    "長打率(パワー)": min(row["長打率"] / 0.600, 1.0),
                    "出塁率(選球眼)": min(row["出塁率"] / 0.450, 1.0),
                    "OPS(貢献度)": min(row["OPS"] / 1.000, 1.0),
                    "打点(勝負強さ)": min(row["打点"] / 15, 1.0)
                }
                
                df_graph = pd.DataFrame(dict(r=list(stats.values()), theta=list(stats.keys())))
                fig = px.line_polar(df_graph, r='r', theta='theta', line_close=True, range_r=[0,1])
                fig.update_traces(fill='toself', line_color="#2E7D32")
                fig.update_layout(polar=dict(radialaxis=dict(visible=True, showticklabels=False)), margin=dict(l=40, r=40, t=40, b=40))
                st.plotly_chart(fig, use_container_width=True)
                
            # --- 投手のレーダーチャート描画 ---
            elif not p_p.empty:
                row = p_p.iloc[0]
                stats_p = {
                    "防御率(安定感)": max(0.0, min(1.0, (6.0 - row["防御率"]) / 6.0)),
                    "WHIP(支配力)": max(0.0, min(1.0, (2.0 - row["WHIP"]) / 2.0)),
                    "与球率(制球力)": max(0.0, min(1.0, (5.0 - row["与球率"]) / 5.0)),
                    "奪三振率(キレ)": min(row["奪三振率"] / 10.0, 1.0)
                }
                df_graph = pd.DataFrame(dict(r=list(stats_p.values()), theta=list(stats_p.keys())))
                fig = px.line_polar(df_graph, r='r', theta='theta', line_close=True, range_r=[0,1])
                fig.update_traces(fill='toself', line_color="#EF6C00")
                fig.update_layout(polar=dict(radialaxis=dict(visible=True, showticklabels=False)), margin=dict(l=40, r=40, t=40, b=40))
                st.plotly_chart(fig, use_container_width=True)